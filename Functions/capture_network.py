import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import asyncio

async def capture_network(page, output_file, testcase_index):
    requests_log = []
    responses_log = []

    async def log_request(request):
        request_start_time = time.time()
        request_data = {
            'method': request.method,
            'url': request.url,
            'headers': request.headers,
            'post_data': request.post_data,
            'start_time': request_start_time,
            'testcase': testcase_index
        }
        requests_log.append(request_data)
        request.start_time = request_start_time

    async def log_response(response):
        request = response.request
        try:
            response_end_time = time.time()
            response_time = response_end_time - request.start_time
            content_type = response.headers.get('content-type', '')

            if 'application/json' in content_type or 'text' in content_type:
                body = await response.text()
            else:
                body = f"Binary data ({response.headers.get('content-length', 'unknown')} bytes)"

            response_data = {
                'url': request.url,
                'status': response.status,
                'headers': response.headers,
                'body': body,
                'response_time_seconds': response_time,
                'response_size_bytes': len(body) if isinstance(body, str) else None,
                'testcase': testcase_index
            }
            responses_log.append(response_data)

        except Exception as e:
            error_data = {
                'url': request.url,
                'status': response.status,
                'headers': response.headers,
                'body': f"Error reading response: {e}",
                'response_time_seconds': None,
                'response_size_bytes': None,
                'testcase': testcase_index
            }
            responses_log.append(error_data)

    page.on("request", log_request)
    page.on("response", log_response)

    await asyncio.sleep(10)  # Ensure some delay to allow all network requests to complete

    # Create and save the Excel workbook
    workbook = Workbook()

    # Create sheets for requests and responses
    requests_sheet = workbook.create_sheet(title="Requests")
    responses_sheet = workbook.create_sheet(title="Responses")

    # Remove the default sheet
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)

    requests_headers = ["Method", "URL", "Headers", "Post Data", "Start Time", "Test Case"]
    responses_headers = ["URL", "Status", "Headers", "Body", "Response Time (seconds)", "Response Size (bytes)", "Test Case"]

    requests_sheet.append(requests_headers)
    responses_sheet.append(responses_headers)

    for request in requests_log:
        row = [
            request['method'], 
            request['url'], 
            str(request['headers']), 
            request['post_data'], 
            time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(request['start_time'])),
            request['testcase']
        ]
        requests_sheet.append(row)

    for response in responses_log:
        row = [
            response['url'], 
            response['status'], 
            str(response['headers']), 
            response['body'], 
            response['response_time_seconds'], 
            response['response_size_bytes'],
            response['testcase']
        ]
        responses_sheet.append(row)

    # Add conditional formatting for response time
    for col in responses_sheet.iter_cols(min_col=5, max_col=5, min_row=2, max_row=len(responses_log) + 1):
        for cell in col:
            if cell.value and cell.value >= 1:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    workbook.save(output_file)
