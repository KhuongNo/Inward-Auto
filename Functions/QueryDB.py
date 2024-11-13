# import cx_Oracle
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import os

# Oracle database connection details
# DSN = cx_Oracle.makedsn("172.29.255.126", "1521", service_name="orcl")
USER = "SAALEM_DEBUG"
PASSWORD = "SAALEM_DEBUG"

# Initialize the Oracle client
# cx_Oracle.init_oracle_client(lib_dir=r"C:\Users\khuongn\Oracle\instantclient_23_4")

# Define the output directory and timestamp globally
output_dir = r'D:\PythonTest\QueryDB'
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

# def query_database(query):
    # Establish the database connection
    # connection = cx_Oracle.connect(user=USER, password=PASSWORD, dsn=DSN)
    # cursor = connection.cursor()
    
    # Execute the query
    # cursor.execute(query)
    
    # Fetch the results
    # columns = [col[0] for col in cursor.description]
    # rows = cursor.fetchall()
    
    # Convert to a DataFrame
    # df = pd.DataFrame(rows, columns=columns)
    
    # Close the cursor and connection
    # cursor.close()
    # connection.close()
    
# return df

def append_df_to_excel(writer, df, sheet_name):
    book = writer.book
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        for r in dataframe_to_rows(df, index=False, header=False):
            sheet.append(r)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def save_to_excel(dataframes, output_file):
    if os.path.exists(output_file):
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            for df, table_name in dataframes:
                append_df_to_excel(writer, df, table_name)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for df, table_name in dataframes:
                df.to_excel(writer, sheet_name=table_name, index=False)
                
    # Ensure at least one sheet is visible
    book = load_workbook(output_file)
    first_sheet = book.sheetnames[0]
    book[first_sheet].sheet_state = 'visible'
    book.save(output_file)

def run_multiple_queries_and_save(queries, index):
    dataframes = []
    # for query, table_name in queries:
        # df = query_database(query)
        # dataframes.append((df, table_name))
    
    os.makedirs(output_dir, exist_ok=True)  # Ensure the directory exists
    output_file = os.path.join(output_dir, f"QueryDB_Testcase_{index}_{timestamp}.xlsx")
    
    save_to_excel(dataframes, output_file)
    print(f"Database log Save to {output_file}")

# Example usage
if __name__ == "__main__":
    extracted_data = "X"  # Replace with the actual extracted data value
    queries = [
        (f"SELECT * FROM saalem.rlos_kt_process WHERE rloskt_code = '{extracted_data}'", "rlos_kt_process"),
        (f"SELECT * FROM saalem.los_link_process WHERE root_code = '{extracted_data}'", "los_link_process")
    ]
    index = 1  # Example index value
    run_multiple_queries_and_save(queries, index)
