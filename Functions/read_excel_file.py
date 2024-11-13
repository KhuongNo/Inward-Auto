import openpyxl

def read_excel_file(file_path: str):
    # Sheets to read from, defined directly in the function
    sheets_to_read = ['BUOC1', 'BUOC2', 'BUOC3', 'BUOC4', 'BUOC5', 'BUOC6']

    def safe_str(value):
        return str(value) if value is not None else ""

    workbook = openpyxl.load_workbook(file_path)
    data = []

    # Ensure both sheets exist in the workbook
    if all(sheet in workbook.sheetnames for sheet in sheets_to_read):
        # Read the data from each sheet
        sheet_data = []
        for sheet_name in sheets_to_read:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Skipping the header row
            sheet_data.append(rows)

        # Combine the data from each row across the sheets
        for i in range(len(sheet_data[0])):  # Assuming both sheets have the same number of rows
            combined_data = {}

            # Extract data from the first sheet (Sheet A)
            for key, value in {
                "chon_gddn": safe_str(sheet_data[0][i][0]),
                "user_buoc1": safe_str(sheet_data[0][i][1]),
                "pass_buoc1": safe_str(sheet_data[0][i][2])
            }.items():
                if key in combined_data:
                    print(f"Duplicate key found: {key}")
                    raise ValueError(f"Duplicate key found: {key}")
                combined_data[key] = value

            # Extract data from the second sheet (Sheet B)
            for key, value in {
                "user_buoc2": safe_str(sheet_data[1][i][0]),
                "pass_buoc2": safe_str(sheet_data[1][i][1])
            }.items():
                if key in combined_data:
                    print(f"Duplicate key found: {key}")
                    raise ValueError(f"Duplicate key found: {key}")
                combined_data[key] = value

            for key, value in {
                "user_buoc3": safe_str(sheet_data[2][i][0]),
                "pass_buoc3": safe_str(sheet_data[2][i][0])
            }.items():
                if key in combined_data:
                    print(f"Duplicate key found: {key}")
                    raise ValueError(f"Duplicate key found: {key}")
                combined_data[key] = value
            data.append(combined_data)

    return data
